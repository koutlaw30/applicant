from tkinter import ttk
import tkinter as tk
from tkinter.scrolledtext import ScrolledText
from tkinter import filedialog as fd
from tkinter import *
from tkinter import messagebox
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
import PyPDF2
import re
import pyperclip
import os
#import win32.com.client as win32


##def  sendemail():
##        outlook = win32Dispatch(
##        'outlook.application')
##        mail = outlook.CreateItem(0)
##        mailing_list = []
##
##        mail.To = 'koutlaw30@gmail.com'
##        mail.Subject = 'Survey of Interest - WSMC - ' + str(myjobtitle.get())  ' - ' + str(myseries.get()) ' - ' + str(mycert.get())
##
##        mail.Body = "Good Morning,"
##        "This email is being sent as a survey of interest to fill the position of "
##        + str(myjobtitle.get()) ' - ' + str(myseries.get()) ", within the Weapon"
##        "Systems Management Center. Your name has been received on a certificate of"
##        "eligible candidates for this position." & vbNewLine & vbNewLine _
##        "Please reply to this email with your response, by " str(mytime.get()) " Eastern Standard"
##        "Time (EST) on " & str(myduedate()) " stating if you wish or do not wish to be further"
##        "considered  for this position. Negative responses are requested."
##        "If you would like verification that your response was received, please use"
##        "the 'Read Receipt' option."
##        "Thank you in advance."
##        "v/r"
##        "Kimberly Hamblen"
##        "Marine Corps Logistics Command"
##        "Weapon Systems Management Center"
##        "Center Operations Division, Albany, GA"
##        "229-639-9958"
##
##        "Kendall"
##        attachment = r"Real path to file."
##
##        mail.Attachments.Add(Source
##                              =attatchment.format(os.getcwd()
##                                ,'file name'))
##        mail.Send()
##        
##        messagebox.showinfo("Send Email", "The email has been sent")





def file_content():
    """ this function extracts content from a PDF file """
    pdf_stream = open('resume.pdf', 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdf_stream)
    first_page_content = pdfReader.getPage(0).extractText().replace('\n', '')
    pdf_stream.close()
    return first_page_content


def email_finder(string_content):
    """ returns emails in a list """
    email_list = []
    email_pattern = r"([\w\.-]+)@([\w\.-]+)(\.[\w\.]+)"
    emails = re.findall(pattern=email_pattern, string=string_content)
    for email_ in emails:
        email = email_[0] + '@' + email_[1] + email_[2]
        email_list.append(email)

    return email_list


pdf_content = file_content()
#print(email_finder(pdf_content))

def deleteselectedapplicants():
         selection = applicantlistbox.curselection()
         applicantlistbox.delete(selection[0])

def deleteallapplicants():
         applicantlistbox.delete(0,END)


def deleteallpanelmembers():
         panellistbox.delete(0,END)


def deleteselectedemails():
         selection = email_list_box.curselection()
         email_list_box.delete(selection[0])

def addemails():
         email_list_box.insert(0, myemail.get())

def addapplicants():
         applicantlistbox.insert(0, myapplicant.get())

         
def scoresheets():
        z = StringVar()
        y = StringVar()
        wb = openpyxl.load_workbook('My Name Application.xlsx')
        #ws2 = wb.create_sheet(str((applicantlistbox.get(applicantlistbox.curselection()))))
        #wb[(str((applicantlistbox.get(applicantlistbox.curselection()))))]['A1'] = "INDIVIDUAL SCORE SHEET"
        for z, listbox_entry in enumerate(applicantlistbox.get(0, END)):
            for y, listbox_entry in enumerate(panellistbox.get(0, END)):
            #If NumberofInterviewQuestions.Value = 6 Then

  


                # new method
                formatted_name = f'''{applicantlistbox.get(z).split(', ')[0]} {applicantlistbox.get(z).split(', ')[1][0]} {panellistbox.get(y).split(' ')[0][0]} {panellistbox.get(y).split(' ')[1][0]}''' + " Score"
                ws2 = wb.create_sheet(formatted_name)
                a1 = wb[formatted_name]['A1']
                ft = wb[formatted_name].font = Font(name='Arial', size=12, bold=True)
                ft1 = wb[formatted_name].font = Font(name='Arial', bold=True)
                a1.font = ft
                a1 = wb[formatted_name]['A1'].alignment = Alignment(horizontal='center')
                a2 = wb[formatted_name]['A2'].alignment = Alignment(horizontal='center')
                a9 = wb[formatted_name]['A9']
                a12 = wb[formatted_name]['A12']
                a15 = wb[formatted_name]['A15']
                a18 = wb[formatted_name]['A18']
                a21 = wb[formatted_name]['A21']
                a9.font = ft1
                a12.font = ft1
                a15.font = ft1
                a18.font = ft1
                a21.font = ft1
                wb[formatted_name]['A1'] = "INDIVIDUAL SCORE SHEET"
                wb[formatted_name]['A2'] = "RPA: " + str(myrpa.get())
                wb[formatted_name]['A4'] = "Applicant’s Name: " + str(
                    (applicantlistbox.get(z)))
                wb[formatted_name]['A6'] = "Panel Member’s Name: " + str(
                    (panellistbox.get(y)))
                wb[formatted_name]['A8'] = "Individual Application Scores"
                wb[formatted_name]['A9'] = "Max- " + str(maxscore.get())
                wb[formatted_name]['A11'] = "A. Experience"
                wb[formatted_name]['A12'] = "Max- " + str(experience.get())
                wb[formatted_name]['A14'] = "B. Education"
                wb[formatted_name]['A15'] = "Max- " + str(education.get())
                wb[formatted_name]['A17'] = "C. Training"
                wb[formatted_name]['A18'] = "Max- " + str(training.get())
                wb[formatted_name]['A20'] = "D. Awards"
                wb[formatted_name]['A21'] = "Max- " + str(awards.get())
                wb[formatted_name]['A24'] = "Total Application Points"
                wb[formatted_name].merge_cells('A1:G1')
                wb[formatted_name].row_dimensions[1].height = 20
                wb[formatted_name].row_dimensions[2].height = 18
                wb[formatted_name].row_dimensions[3].height = 18
                wb[formatted_name].row_dimensions[4].height = 15
                wb[formatted_name].row_dimensions[5 - 57].height = 15.75
                wb[formatted_name].column_dimensions['A'].width = 30.86
                wb[formatted_name].column_dimensions['B'].width = 6.57
                wb[formatted_name].column_dimensions['C'].width = 8.43
                wb[formatted_name].column_dimensions['D'].width = 8.57
                wb[formatted_name].column_dimensions['E'].width = 15.29
                wb[formatted_name].column_dimensions['F'].width = 19
                wb[formatted_name].column_dimensions['G'].width = 21.7
                thin_border = Border(bottom=Side(style='thin')) 
                wb[formatted_name]['A9'].border = thin_border
                wb[formatted_name]['A12'].border = thin_border
                wb[formatted_name]['A15'].border = thin_border
                wb[formatted_name]['A18'].border = thin_border
                wb[formatted_name]['A21'].border = thin_border
                wb[formatted_name].page_setup.PrintArea = ""
                wb[formatted_name].page_setup.LeftHeader = ""
                wb[formatted_name].page_setup.CenterHeader = ""
                wb[formatted_name].page_setup.RightHeader = ""
                wb[formatted_name].page_setup.LeftFooter = ""
                wb[formatted_name].page_setup.CenterFooter = ""
                wb[formatted_name].page_setup.RightFooter = ""
                wb[formatted_name].page_setup.FitToPagesWide = 1
                wb[formatted_name].page_setup.FitToPagesTall = 1
                wb[formatted_name].page_setup.LeftMargin = 0.75
                wb[formatted_name].page_setup.RightMargin = 0.75
                wb[formatted_name].page_setup.TopMargin = .25
                wb[formatted_name].page_setup.BottomMargin = .25
                wb[formatted_name].page_setup.HeaderMargin = 0.5
                wb[formatted_name].page_setup.FooterMargin = 0.5
                wb[formatted_name].page_setup.fitToPage = True
        wb.save('My Name Application.xlsx')
        messagebox.showinfo("Update Spreadsheet", "The spreadsheet has been updated")


def WorksheetNotes():
        z = StringVar()
        y = StringVar()
        wb = openpyxl.load_workbook('My Name Application.xlsx')
        for z, listbox_entry in enumerate(applicantlistbox.get(0, END)):
            for y, listbox_entry in enumerate(panellistbox.get(0, END)):
                formatted_name = f'''{applicantlistbox.get(z).split(', ')[0]} {applicantlistbox.get(z).split(', ')[1][0]} {panellistbox.get(y).split(' ')[0][0]} {panellistbox.get(y).split(' ')[1][0]}''' + " S Notes"
                ws2 = wb.create_sheet(formatted_name)
                wb[formatted_name]['A1'] = "INDIVIDUAL SCORE SHEET NOTES"
                wb[formatted_name]['A2'] = "RPA: " + str(myrpa.get()) + str(myjobtitle.get()) + " - " + str(myseries.get())
                wb[formatted_name]['A4'] = "Applicant’s Name: " + str(
                    (applicantlistbox.get(z)))
                wb[formatted_name]['A5'] = "Panel Member’s Name: " + str(
                    (panellistbox.get(y)))
                wb[formatted_name]['A7'] = "Experience:"
                wb[formatted_name]['A27'] = "Education:"
                wb[formatted_name]['A31'] = "Training:"
                wb[formatted_name]['A35'] = "Awards:"
                wb[formatted_name]['A39'] = "Please do not write on the back. Only use paper that is provided – NOT notebook paper. If additional paper is required, contact admin POC. Thank you!"
                wb[formatted_name].row_dimensions[1].height = 19.5
                wb[formatted_name].row_dimensions[2].height = 17.25
                wb[formatted_name].row_dimensions[3].height = 12.75
                wb[formatted_name].row_dimensions[4].height = 15.75
                wb[formatted_name].row_dimensions[5].height = 15.75
                wb[formatted_name].row_dimensions[6].height = 15.75
                wb[formatted_name].row_dimensions[7].height = 16.75
                wb[formatted_name].row_dimensions[8].height = 16.75
                wb[formatted_name].row_dimensions[9].height = 16.75
                wb[formatted_name].row_dimensions[10].height = 16.75
                wb[formatted_name].row_dimensions[11].height = 16.75
                wb[formatted_name].row_dimensions[12].height = 16.75
                wb[formatted_name].row_dimensions[13].height = 16.75
                wb[formatted_name].row_dimensions[14].height = 16.75
                wb[formatted_name].row_dimensions[15].height = 16.75
                wb[formatted_name].row_dimensions[16].height = 16.75
                wb[formatted_name].row_dimensions[17].height = 16.75
                wb[formatted_name].row_dimensions[18].height = 16.75
                wb[formatted_name].row_dimensions[19].height = 16.75
                wb[formatted_name].row_dimensions[20].height = 16.75
                wb[formatted_name].row_dimensions[21].height = 16.75
                wb[formatted_name].row_dimensions[22].height = 16.75
                wb[formatted_name].row_dimensions[23].height = 16.75
                wb[formatted_name].row_dimensions[24].height = 16.75
                wb[formatted_name].row_dimensions[25].height = 16.75
                wb[formatted_name].row_dimensions[26].height = 16.75
                wb[formatted_name].row_dimensions[27].height = 16.75
                wb[formatted_name].row_dimensions[28].height = 16.75
                wb[formatted_name].row_dimensions[29].height = 16.75
                wb[formatted_name].row_dimensions[30].height = 16.75
                wb[formatted_name].row_dimensions[31].height = 16.75
                wb[formatted_name].row_dimensions[32].height = 16.75
                wb[formatted_name].row_dimensions[33].height = 16.75
                wb[formatted_name].row_dimensions[34].height = 16.75
                wb[formatted_name].row_dimensions[35].height = 16.75
                wb[formatted_name].row_dimensions[36].height = 16.75            
                wb[formatted_name].row_dimensions[37].height = 16.75
                wb[formatted_name].row_dimensions[38].height = 16.75
                wb[formatted_name].row_dimensions[39].height = 47.25
                wb[formatted_name].column_dimensions['A'].width = 20
                wb[formatted_name].column_dimensions['B'].width = 5.86
                wb[formatted_name].column_dimensions['C'].width = 7.71
                wb[formatted_name].column_dimensions['D'].width = 7.86
                wb[formatted_name].column_dimensions['E'].width = 9.86
                wb[formatted_name].column_dimensions['F'].width = 10.14
                wb[formatted_name].column_dimensions['G'].width = 27
                a1 = wb[formatted_name]['A1']
                ft = wb[formatted_name].font = Font(name='Calibri', size=14, bold=True)
                ft1 = wb[formatted_name].font = Font(name='Arial', size=12, bold=True)
                a1.font = ft
                a1 = wb[formatted_name]['A1']
                wb[formatted_name].merge_cells('A1:G1')
                wb[formatted_name].merge_cells('A2:G2')
                wb[formatted_name].merge_cells('A39:G39')
                a2 = wb[formatted_name]['A2']
                a2.font = ft
                a4 = wb[formatted_name]['A4']
                a5 = wb[formatted_name]['A5']
                b4 = wb[formatted_name]['B4']
                c4 = wb[formatted_name]['C4']
                d4 = wb[formatted_name]['D4']
                e4 = wb[formatted_name]['E4']
                f4 = wb[formatted_name]['F4']
                g4 = wb[formatted_name]['G4']
                a7 = wb[formatted_name]['A7']
                a8 = wb[formatted_name]['A8']
                a9 = wb[formatted_name]['A9']
                a10 = wb[formatted_name]['A10']
                a11 = wb[formatted_name]['A11']
                a12 = wb[formatted_name]['A12']
                a13 = wb[formatted_name]['A13']
                a14 = wb[formatted_name]['A14']
                a15 = wb[formatted_name]['A15']
                a16 = wb[formatted_name]['A16']
                a17 = wb[formatted_name]['A17']
                a18 = wb[formatted_name]['A18']
                a19 = wb[formatted_name]['A19']
                a20 = wb[formatted_name]['A20']
                a21 = wb[formatted_name]['A21']
                a22 = wb[formatted_name]['A22']
                a23 = wb[formatted_name]['A23']
                a24 = wb[formatted_name]['A24']
                a25 = wb[formatted_name]['A25']
                a26 = wb[formatted_name]['A26']
                a27 = wb[formatted_name]['A27']
                a28 = wb[formatted_name]['A28']
                a29 = wb[formatted_name]['A29']
                a30 = wb[formatted_name]['A30']
                a31 = wb[formatted_name]['A31']
                a32 = wb[formatted_name]['A32']
                a33 = wb[formatted_name]['A33']
                a34 = wb[formatted_name]['A34']
                a35 = wb[formatted_name]['A35']
                a1 = wb[formatted_name]['A1'].alignment = Alignment(horizontal='center')
                a2 = wb[formatted_name]['A2'].alignment = Alignment(horizontal='center')
                a4.font = ft
                a5.font = ft
                b4.font = ft
                c4.font = ft
                d4.font = ft
                e4.font = ft
                f4.font = ft
                g4.font = ft
                a7.font = ft1
                a8.font = ft1
                a9.font = ft1
                a10.font = ft1
                a11.font = ft1
                a12.font = ft1
                a13.font = ft1
                a14.font = ft1
                a15.font = ft1
                a16.font = ft1
                a17.font = ft1
                a18.font = ft1
                a19.font = ft1
                a20.font = ft1
                a21.font = ft1
                a22.font = ft1
                a23.font = ft1
                a24.font = ft1
                a25.font = ft1
                a26.font = ft1
                a27.font = ft1
                a28.font = ft1
                a29.font = ft1
                a30.font = ft1
                a31.font = ft1
                a32.font = ft1
                a33.font = ft1
                a34.font = ft1
                a35.font = ft1
                thin_border = Border(bottom=Side(style='thin')) 
                wb[formatted_name]['A7'].border = thin_border
                wb[formatted_name]['B7'].border = thin_border
                wb[formatted_name]['C7'].border = thin_border
                wb[formatted_name]['D7'].border = thin_border
                wb[formatted_name]['E7'].border = thin_border
                wb[formatted_name]['F7'].border = thin_border
                wb[formatted_name]['G7'].border = thin_border
                wb[formatted_name]['A8'].border = thin_border
                wb[formatted_name]['B8'].border = thin_border
                wb[formatted_name]['C8'].border = thin_border
                wb[formatted_name]['D8'].border = thin_border
                wb[formatted_name]['E8'].border = thin_border
                wb[formatted_name]['F8'].border = thin_border
                wb[formatted_name]['G8'].border = thin_border
                wb[formatted_name]['A9'].border = thin_border
                wb[formatted_name]['B9'].border = thin_border
                wb[formatted_name]['C9'].border = thin_border
                wb[formatted_name]['D9'].border = thin_border
                wb[formatted_name]['E9'].border = thin_border
                wb[formatted_name]['F9'].border = thin_border
                wb[formatted_name]['G9'].border = thin_border
                wb[formatted_name]['A10'].border = thin_border
                wb[formatted_name]['B10'].border = thin_border
                wb[formatted_name]['C10'].border = thin_border
                wb[formatted_name]['D10'].border = thin_border
                wb[formatted_name]['E10'].border = thin_border
                wb[formatted_name]['F10'].border = thin_border
                wb[formatted_name]['G10'].border = thin_border
                wb[formatted_name]['A11'].border = thin_border
                wb[formatted_name]['B11'].border = thin_border
                wb[formatted_name]['C11'].border = thin_border
                wb[formatted_name]['D11'].border = thin_border
                wb[formatted_name]['E11'].border = thin_border
                wb[formatted_name]['F11'].border = thin_border
                wb[formatted_name]['G11'].border = thin_border
                wb[formatted_name]['A12'].border = thin_border
                wb[formatted_name]['B12'].border = thin_border
                wb[formatted_name]['C12'].border = thin_border
                wb[formatted_name]['D12'].border = thin_border
                wb[formatted_name]['E12'].border = thin_border
                wb[formatted_name]['F12'].border = thin_border
                wb[formatted_name]['G12'].border = thin_border
                wb[formatted_name]['A13'].border = thin_border
                wb[formatted_name]['B13'].border = thin_border
                wb[formatted_name]['C13'].border = thin_border
                wb[formatted_name]['D13'].border = thin_border
                wb[formatted_name]['E13'].border = thin_border
                wb[formatted_name]['F13'].border = thin_border
                wb[formatted_name]['G13'].border = thin_border
                wb[formatted_name]['A14'].border = thin_border
                wb[formatted_name]['B14'].border = thin_border
                wb[formatted_name]['C14'].border = thin_border
                wb[formatted_name]['D14'].border = thin_border
                wb[formatted_name]['E14'].border = thin_border
                wb[formatted_name]['F14'].border = thin_border
                wb[formatted_name]['G14'].border = thin_border
                wb[formatted_name]['A15'].border = thin_border
                wb[formatted_name]['B15'].border = thin_border
                wb[formatted_name]['C15'].border = thin_border
                wb[formatted_name]['D15'].border = thin_border
                wb[formatted_name]['E15'].border = thin_border
                wb[formatted_name]['F15'].border = thin_border
                wb[formatted_name]['G15'].border = thin_border
                wb[formatted_name]['A16'].border = thin_border
                wb[formatted_name]['B16'].border = thin_border
                wb[formatted_name]['C16'].border = thin_border
                wb[formatted_name]['D16'].border = thin_border
                wb[formatted_name]['E16'].border = thin_border
                wb[formatted_name]['F16'].border = thin_border
                wb[formatted_name]['G16'].border = thin_border
                wb[formatted_name]['A17'].border = thin_border
                wb[formatted_name]['B17'].border = thin_border
                wb[formatted_name]['C17'].border = thin_border
                wb[formatted_name]['D17'].border = thin_border
                wb[formatted_name]['E17'].border = thin_border
                wb[formatted_name]['F17'].border = thin_border
                wb[formatted_name]['G17'].border = thin_border
                wb[formatted_name]['A18'].border = thin_border
                wb[formatted_name]['B18'].border = thin_border
                wb[formatted_name]['C18'].border = thin_border
                wb[formatted_name]['D18'].border = thin_border
                wb[formatted_name]['E18'].border = thin_border
                wb[formatted_name]['F18'].border = thin_border
                wb[formatted_name]['G18'].border = thin_border
                wb[formatted_name]['A19'].border = thin_border
                wb[formatted_name]['B19'].border = thin_border
                wb[formatted_name]['C19'].border = thin_border
                wb[formatted_name]['D19'].border = thin_border
                wb[formatted_name]['E19'].border = thin_border
                wb[formatted_name]['F19'].border = thin_border
                wb[formatted_name]['G19'].border = thin_border
                wb[formatted_name]['A20'].border = thin_border
                wb[formatted_name]['B20'].border = thin_border
                wb[formatted_name]['C20'].border = thin_border
                wb[formatted_name]['D20'].border = thin_border
                wb[formatted_name]['E20'].border = thin_border
                wb[formatted_name]['F20'].border = thin_border
                wb[formatted_name]['G20'].border = thin_border
                wb[formatted_name]['A21'].border = thin_border
                wb[formatted_name]['B21'].border = thin_border
                wb[formatted_name]['C21'].border = thin_border
                wb[formatted_name]['D21'].border = thin_border
                wb[formatted_name]['E21'].border = thin_border
                wb[formatted_name]['F21'].border = thin_border
                wb[formatted_name]['G21'].border = thin_border
                wb[formatted_name]['A22'].border = thin_border
                wb[formatted_name]['B22'].border = thin_border
                wb[formatted_name]['C22'].border = thin_border
                wb[formatted_name]['D22'].border = thin_border
                wb[formatted_name]['E22'].border = thin_border
                wb[formatted_name]['F22'].border = thin_border
                wb[formatted_name]['G22'].border = thin_border
                wb[formatted_name]['A23'].border = thin_border
                wb[formatted_name]['B23'].border = thin_border
                wb[formatted_name]['C23'].border = thin_border
                wb[formatted_name]['D23'].border = thin_border
                wb[formatted_name]['E23'].border = thin_border
                wb[formatted_name]['F23'].border = thin_border
                wb[formatted_name]['G23'].border = thin_border
                wb[formatted_name]['A24'].border = thin_border
                wb[formatted_name]['B24'].border = thin_border
                wb[formatted_name]['C24'].border = thin_border
                wb[formatted_name]['D24'].border = thin_border
                wb[formatted_name]['E24'].border = thin_border
                wb[formatted_name]['F24'].border = thin_border
                wb[formatted_name]['G24'].border = thin_border
                wb[formatted_name]['A25'].border = thin_border
                wb[formatted_name]['B25'].border = thin_border
                wb[formatted_name]['C25'].border = thin_border
                wb[formatted_name]['D25'].border = thin_border
                wb[formatted_name]['E25'].border = thin_border
                wb[formatted_name]['F25'].border = thin_border
                wb[formatted_name]['G25'].border = thin_border
                wb[formatted_name]['A26'].border = thin_border
                wb[formatted_name]['B26'].border = thin_border
                wb[formatted_name]['C26'].border = thin_border
                wb[formatted_name]['D26'].border = thin_border
                wb[formatted_name]['E26'].border = thin_border
                wb[formatted_name]['F26'].border = thin_border
                wb[formatted_name]['G26'].border = thin_border
                wb[formatted_name]['A27'].border = thin_border
                wb[formatted_name]['B27'].border = thin_border
                wb[formatted_name]['C27'].border = thin_border
                wb[formatted_name]['D27'].border = thin_border
                wb[formatted_name]['E27'].border = thin_border
                wb[formatted_name]['F27'].border = thin_border
                wb[formatted_name]['G27'].border = thin_border
                wb[formatted_name]['A28'].border = thin_border
                wb[formatted_name]['B28'].border = thin_border
                wb[formatted_name]['C28'].border = thin_border
                wb[formatted_name]['D28'].border = thin_border
                wb[formatted_name]['E28'].border = thin_border
                wb[formatted_name]['F28'].border = thin_border
                wb[formatted_name]['G28'].border = thin_border
                wb[formatted_name]['A29'].border = thin_border
                wb[formatted_name]['B29'].border = thin_border
                wb[formatted_name]['C29'].border = thin_border
                wb[formatted_name]['D29'].border = thin_border
                wb[formatted_name]['E29'].border = thin_border
                wb[formatted_name]['F29'].border = thin_border
                wb[formatted_name]['G29'].border = thin_border
                wb[formatted_name]['A30'].border = thin_border
                wb[formatted_name]['B30'].border = thin_border
                wb[formatted_name]['C30'].border = thin_border
                wb[formatted_name]['D30'].border = thin_border
                wb[formatted_name]['E30'].border = thin_border
                wb[formatted_name]['F30'].border = thin_border
                wb[formatted_name]['G30'].border = thin_border
                wb[formatted_name]['A31'].border = thin_border
                wb[formatted_name]['B31'].border = thin_border
                wb[formatted_name]['C31'].border = thin_border
                wb[formatted_name]['D31'].border = thin_border
                wb[formatted_name]['E31'].border = thin_border
                wb[formatted_name]['F31'].border = thin_border
                wb[formatted_name]['G31'].border = thin_border
                wb[formatted_name]['A32'].border = thin_border
                wb[formatted_name]['B32'].border = thin_border
                wb[formatted_name]['C32'].border = thin_border
                wb[formatted_name]['D32'].border = thin_border
                wb[formatted_name]['E32'].border = thin_border
                wb[formatted_name]['F32'].border = thin_border
                wb[formatted_name]['G32'].border = thin_border
                wb[formatted_name]['A33'].border = thin_border
                wb[formatted_name]['B33'].border = thin_border
                wb[formatted_name]['C33'].border = thin_border
                wb[formatted_name]['D33'].border = thin_border
                wb[formatted_name]['E33'].border = thin_border
                wb[formatted_name]['F33'].border = thin_border
                wb[formatted_name]['G33'].border = thin_border
                wb[formatted_name]['A34'].border = thin_border
                wb[formatted_name]['B34'].border = thin_border
                wb[formatted_name]['C34'].border = thin_border
                wb[formatted_name]['D34'].border = thin_border
                wb[formatted_name]['E34'].border = thin_border
                wb[formatted_name]['F34'].border = thin_border
                wb[formatted_name]['G34'].border = thin_border
                wb[formatted_name]['A35'].border = thin_border
                wb[formatted_name]['B35'].border = thin_border
                wb[formatted_name]['C35'].border = thin_border
                wb[formatted_name]['D35'].border = thin_border
                wb[formatted_name]['E35'].border = thin_border
                wb[formatted_name]['F35'].border = thin_border
                wb[formatted_name]['G35'].border = thin_border
                wb[formatted_name]['A36'].border = thin_border
                wb[formatted_name]['B36'].border = thin_border
                wb[formatted_name]['C36'].border = thin_border
                wb[formatted_name]['D36'].border = thin_border
                wb[formatted_name]['E36'].border = thin_border
                wb[formatted_name]['F36'].border = thin_border
                wb[formatted_name]['G36'].border = thin_border
                wb[formatted_name]['A37'].border = thin_border
                wb[formatted_name]['B37'].border = thin_border
                wb[formatted_name]['C37'].border = thin_border
                wb[formatted_name]['D37'].border = thin_border
                wb[formatted_name]['E37'].border = thin_border
                wb[formatted_name]['F37'].border = thin_border
                wb[formatted_name]['G37'].border = thin_border
                wb[formatted_name]['A38'].border = thin_border
                wb[formatted_name]['B38'].border = thin_border
                wb[formatted_name]['C38'].border = thin_border
                wb[formatted_name]['D38'].border = thin_border
                wb[formatted_name]['E38'].border = thin_border
                wb[formatted_name]['F38'].border = thin_border
                wb[formatted_name]['G38'].border = thin_border
                a39 = wb[formatted_name]['A39'].alignment = Alignment(horizontal='center')
                a39 = wb[formatted_name]['A39'].alignment = Alignment(wrapText=True)
                wb[formatted_name].page_setup.PrintArea = ""
                wb[formatted_name].page_setup.LeftHeader = ""
                wb[formatted_name].page_setup.CenterHeader = ""
                wb[formatted_name].page_setup.RightHeader = ""
                wb[formatted_name].page_setup.LeftFooter = ""
                wb[formatted_name].page_setup.CenterFooter = ""
                wb[formatted_name].page_setup.RightFooter = ""
                wb[formatted_name].page_setup.FitToPagesWide = .5
                wb[formatted_name].page_setup.FitToPagesTall = .5
                wb[formatted_name].page_setup.CenterHorizontally = True
                wb[formatted_name].page_setup.CenterVertically = True
                wb[formatted_name].page_setup.LeftMargin = 0.75
                wb[formatted_name].page_setup.RightMargin = 0.75
                wb[formatted_name].page_setup.TopMargin = .25
                wb[formatted_name].page_setup.BottomMargin = .25
                wb[formatted_name].page_setup.HeaderMargin = 0.5
                wb[formatted_name].page_setup.FooterMargin = 0.5
                wb[formatted_name].page_setup.fitToPage = True
        wb.save('My Name Application.xlsx')
        messagebox.showinfo("Update Spreadsheet", "The spreadsheet has been updated")


def InterviewNotes():
        z = StringVar()
        y = StringVar()
        wb = openpyxl.load_workbook('My Name Application.xlsx')
        for z, listbox_entry in enumerate(applicantlistbox.get(0, END)):
            for y, listbox_entry in enumerate(panellistbox.get(0, END)):
                formatted_name = f'''{applicantlistbox.get(z).split(', ')[0]} {applicantlistbox.get(z).split(', ')[1][0]} {panellistbox.get(y).split(' ')[0][0]} {panellistbox.get(y).split(' ')[1][0]}''' + " I Notes"
                ws2 = wb.create_sheet(formatted_name)
                wb[formatted_name]['A1'] = "INDIVIDUAL INTERVIEW NOTES"
                wb[formatted_name]['A2'] = "RPA: " + str(myrpa.get()) + str(myjobtitle.get()) + " - " + str(myseries.get())
                wb[formatted_name]['A4'] = "Applicant’s Name: " + str(
                    (applicantlistbox.get(z)))
                wb[formatted_name]['A5'] = "Panel Member’s Name: " + str(
                    (panellistbox.get(y)))
                wb[formatted_name]['A7'] = "Question 1:"
                wb[formatted_name]['A12'] = "Question 2:"
                wb[formatted_name]['A17'] = "Question 3:"
                wb[formatted_name]['A22'] = "Question 4:"
                wb[formatted_name]['A27'] = "Question 5:"
                wb[formatted_name]['A32'] = "Question 6:"
                wb[formatted_name]['A39'] = "Keep with Resume, Individual Score Sheet & Individual Score Sheet Notes. Please do not write on the back. If additional paper is needed, only use printer paper!"
                wb[formatted_name].row_dimensions[1].height = 19.5
                wb[formatted_name].row_dimensions[2].height = 17.25
                wb[formatted_name].row_dimensions[3].height = 12.75
                wb[formatted_name].row_dimensions[4].height = 15.75
                wb[formatted_name].row_dimensions[5].height = 15.75
                wb[formatted_name].row_dimensions[6].height = 15.75
                wb[formatted_name].row_dimensions[39].height = 47.25
                wb[formatted_name].column_dimensions['A'].width = 17
                wb[formatted_name].column_dimensions['B'].width = 10
                wb[formatted_name].column_dimensions['C'].width = 10
                wb[formatted_name].column_dimensions['D'].width = 10
                wb[formatted_name].column_dimensions['E'].width = 10
                wb[formatted_name].column_dimensions['F'].width = 10
                wb[formatted_name].column_dimensions['G'].width = 28.75
                a1 = wb[formatted_name]['A1']
                ft = wb[formatted_name].font = Font(name='Calibri', size=14, bold=True)
                ft1 = wb[formatted_name].font = Font(name='Arial', size=12, bold=True)
                a1.font = ft
                a1 = wb[formatted_name]['A1']
                wb[formatted_name].merge_cells('A1:G1')
                wb[formatted_name].merge_cells('A2:G2')
                wb[formatted_name].merge_cells('A39:G39')
                wb[formatted_name].merge_cells('A2:G2')
                wb[formatted_name].merge_cells('A39:G39')
                a2 = wb[formatted_name]['A2']
                a2.font = ft1
                a4 = wb[formatted_name]['A4']
                a5 = wb[formatted_name]['A5']
                b4 = wb[formatted_name]['B4']
                c4 = wb[formatted_name]['C4']
                d4 = wb[formatted_name]['D4']
                e4 = wb[formatted_name]['E4']
                f4 = wb[formatted_name]['F4']
                g4 = wb[formatted_name]['G4']
                a7 = wb[formatted_name]['A7']
                a8 = wb[formatted_name]['A8']
                a9 = wb[formatted_name]['A9']
                a10 = wb[formatted_name]['A10']
                a11 = wb[formatted_name]['A11']
                a12 = wb[formatted_name]['A12']
                a13 = wb[formatted_name]['A13']
                a14 = wb[formatted_name]['A14']
                a15 = wb[formatted_name]['A15']
                a16 = wb[formatted_name]['A16']
                a17 = wb[formatted_name]['A17']
                a18 = wb[formatted_name]['A18']
                a19 = wb[formatted_name]['A19']
                a20 = wb[formatted_name]['A20']
                a21 = wb[formatted_name]['A21']
                a22 = wb[formatted_name]['A22']
                a23 = wb[formatted_name]['A23']
                a24 = wb[formatted_name]['A24']
                a25 = wb[formatted_name]['A25']
                a26 = wb[formatted_name]['A26']
                a27 = wb[formatted_name]['A27']
                a28 = wb[formatted_name]['A28']
                a29 = wb[formatted_name]['A29']
                a30 = wb[formatted_name]['A30']
                a31 = wb[formatted_name]['A31']
                a32 = wb[formatted_name]['A32']
                a33 = wb[formatted_name]['A33']
                a34 = wb[formatted_name]['A34']
                a35 = wb[formatted_name]['A35']
                a1 = wb[formatted_name]['A1'].alignment = Alignment(horizontal='center')
                a2 = wb[formatted_name]['A2'].alignment = Alignment(horizontal='center')
                a4.font = ft
                a5.font = ft
                b4.font = ft
                c4.font = ft
                d4.font = ft
                e4.font = ft
                f4.font = ft
                g4.font = ft
                a7.font = ft1
                a8.font = ft1
                a9.font = ft1
                a10.font = ft1
                a11.font = ft1
                a12.font = ft1
                a13.font = ft1
                a14.font = ft1
                a15.font = ft1
                a16.font = ft1
                a17.font = ft1
                a18.font = ft1
                a19.font = ft1
                a20.font = ft1
                a21.font = ft1
                a22.font = ft1
                a23.font = ft1
                a24.font = ft1
                a25.font = ft1
                a26.font = ft1
                a27.font = ft1
                a28.font = ft1
                a29.font = ft1
                a30.font = ft1
                a31.font = ft1
                a32.font = ft1
                a33.font = ft1
                a34.font = ft1
                a35.font = ft1
                thin_border = Border(bottom=Side(style='thin')) 
                wb[formatted_name]['A7'].border = thin_border
                wb[formatted_name]['B7'].border = thin_border
                wb[formatted_name]['C7'].border = thin_border
                wb[formatted_name]['D7'].border = thin_border
                wb[formatted_name]['E7'].border = thin_border
                wb[formatted_name]['F7'].border = thin_border
                wb[formatted_name]['G7'].border = thin_border
                wb[formatted_name]['A8'].border = thin_border
                wb[formatted_name]['B8'].border = thin_border
                wb[formatted_name]['C8'].border = thin_border
                wb[formatted_name]['D8'].border = thin_border
                wb[formatted_name]['E8'].border = thin_border
                wb[formatted_name]['F8'].border = thin_border
                wb[formatted_name]['G8'].border = thin_border
                wb[formatted_name]['A9'].border = thin_border
                wb[formatted_name]['B9'].border = thin_border
                wb[formatted_name]['C9'].border = thin_border
                wb[formatted_name]['D9'].border = thin_border
                wb[formatted_name]['E9'].border = thin_border
                wb[formatted_name]['F9'].border = thin_border
                wb[formatted_name]['G9'].border = thin_border
                wb[formatted_name]['A10'].border = thin_border
                wb[formatted_name]['B10'].border = thin_border
                wb[formatted_name]['C10'].border = thin_border
                wb[formatted_name]['D10'].border = thin_border
                wb[formatted_name]['E10'].border = thin_border
                wb[formatted_name]['F10'].border = thin_border
                wb[formatted_name]['G10'].border = thin_border
                wb[formatted_name]['A11'].border = thin_border
                wb[formatted_name]['B11'].border = thin_border
                wb[formatted_name]['C11'].border = thin_border
                wb[formatted_name]['D11'].border = thin_border
                wb[formatted_name]['E11'].border = thin_border
                wb[formatted_name]['F11'].border = thin_border
                wb[formatted_name]['G11'].border = thin_border
                wb[formatted_name]['A12'].border = thin_border
                wb[formatted_name]['B12'].border = thin_border
                wb[formatted_name]['C12'].border = thin_border
                wb[formatted_name]['D12'].border = thin_border
                wb[formatted_name]['E12'].border = thin_border
                wb[formatted_name]['F12'].border = thin_border
                wb[formatted_name]['G12'].border = thin_border
                wb[formatted_name]['A13'].border = thin_border
                wb[formatted_name]['B13'].border = thin_border
                wb[formatted_name]['C13'].border = thin_border
                wb[formatted_name]['D13'].border = thin_border
                wb[formatted_name]['E13'].border = thin_border
                wb[formatted_name]['F13'].border = thin_border
                wb[formatted_name]['G13'].border = thin_border
                wb[formatted_name]['A14'].border = thin_border
                wb[formatted_name]['B14'].border = thin_border
                wb[formatted_name]['C14'].border = thin_border
                wb[formatted_name]['D14'].border = thin_border
                wb[formatted_name]['E14'].border = thin_border
                wb[formatted_name]['F14'].border = thin_border
                wb[formatted_name]['G14'].border = thin_border
                wb[formatted_name]['A15'].border = thin_border
                wb[formatted_name]['B15'].border = thin_border
                wb[formatted_name]['C15'].border = thin_border
                wb[formatted_name]['D15'].border = thin_border
                wb[formatted_name]['E15'].border = thin_border
                wb[formatted_name]['F15'].border = thin_border
                wb[formatted_name]['G15'].border = thin_border
                wb[formatted_name]['A16'].border = thin_border
                wb[formatted_name]['B16'].border = thin_border
                wb[formatted_name]['C16'].border = thin_border
                wb[formatted_name]['D16'].border = thin_border
                wb[formatted_name]['E16'].border = thin_border
                wb[formatted_name]['F16'].border = thin_border
                wb[formatted_name]['G16'].border = thin_border
                wb[formatted_name]['A17'].border = thin_border
                wb[formatted_name]['B17'].border = thin_border
                wb[formatted_name]['C17'].border = thin_border
                wb[formatted_name]['D17'].border = thin_border
                wb[formatted_name]['E17'].border = thin_border
                wb[formatted_name]['F17'].border = thin_border
                wb[formatted_name]['G17'].border = thin_border
                wb[formatted_name]['A18'].border = thin_border
                wb[formatted_name]['B18'].border = thin_border
                wb[formatted_name]['C18'].border = thin_border
                wb[formatted_name]['D18'].border = thin_border
                wb[formatted_name]['E18'].border = thin_border
                wb[formatted_name]['F18'].border = thin_border
                wb[formatted_name]['G18'].border = thin_border
                wb[formatted_name]['A19'].border = thin_border
                wb[formatted_name]['B19'].border = thin_border
                wb[formatted_name]['C19'].border = thin_border
                wb[formatted_name]['D19'].border = thin_border
                wb[formatted_name]['E19'].border = thin_border
                wb[formatted_name]['F19'].border = thin_border
                wb[formatted_name]['G19'].border = thin_border
                wb[formatted_name]['A20'].border = thin_border
                wb[formatted_name]['B20'].border = thin_border
                wb[formatted_name]['C20'].border = thin_border
                wb[formatted_name]['D20'].border = thin_border
                wb[formatted_name]['E20'].border = thin_border
                wb[formatted_name]['F20'].border = thin_border
                wb[formatted_name]['G20'].border = thin_border
                wb[formatted_name]['A21'].border = thin_border
                wb[formatted_name]['B21'].border = thin_border
                wb[formatted_name]['C21'].border = thin_border
                wb[formatted_name]['D21'].border = thin_border
                wb[formatted_name]['E21'].border = thin_border
                wb[formatted_name]['F21'].border = thin_border
                wb[formatted_name]['G21'].border = thin_border
                wb[formatted_name]['A22'].border = thin_border
                wb[formatted_name]['B22'].border = thin_border
                wb[formatted_name]['C22'].border = thin_border
                wb[formatted_name]['D22'].border = thin_border
                wb[formatted_name]['E22'].border = thin_border
                wb[formatted_name]['F22'].border = thin_border
                wb[formatted_name]['G22'].border = thin_border
                wb[formatted_name]['A23'].border = thin_border
                wb[formatted_name]['B23'].border = thin_border
                wb[formatted_name]['C23'].border = thin_border
                wb[formatted_name]['D23'].border = thin_border
                wb[formatted_name]['E23'].border = thin_border
                wb[formatted_name]['F23'].border = thin_border
                wb[formatted_name]['G23'].border = thin_border
                wb[formatted_name]['A24'].border = thin_border
                wb[formatted_name]['B24'].border = thin_border
                wb[formatted_name]['C24'].border = thin_border
                wb[formatted_name]['D24'].border = thin_border
                wb[formatted_name]['E24'].border = thin_border
                wb[formatted_name]['F24'].border = thin_border
                wb[formatted_name]['G24'].border = thin_border
                wb[formatted_name]['A25'].border = thin_border
                wb[formatted_name]['B25'].border = thin_border
                wb[formatted_name]['C25'].border = thin_border
                wb[formatted_name]['D25'].border = thin_border
                wb[formatted_name]['E25'].border = thin_border
                wb[formatted_name]['F25'].border = thin_border
                wb[formatted_name]['G25'].border = thin_border
                wb[formatted_name]['A26'].border = thin_border
                wb[formatted_name]['B26'].border = thin_border
                wb[formatted_name]['C26'].border = thin_border
                wb[formatted_name]['D26'].border = thin_border
                wb[formatted_name]['E26'].border = thin_border
                wb[formatted_name]['F26'].border = thin_border
                wb[formatted_name]['G26'].border = thin_border
                wb[formatted_name]['A27'].border = thin_border
                wb[formatted_name]['B27'].border = thin_border
                wb[formatted_name]['C27'].border = thin_border
                wb[formatted_name]['D27'].border = thin_border
                wb[formatted_name]['E27'].border = thin_border
                wb[formatted_name]['F27'].border = thin_border
                wb[formatted_name]['G27'].border = thin_border
                wb[formatted_name]['A28'].border = thin_border
                wb[formatted_name]['B28'].border = thin_border
                wb[formatted_name]['C28'].border = thin_border
                wb[formatted_name]['D28'].border = thin_border
                wb[formatted_name]['E28'].border = thin_border
                wb[formatted_name]['F28'].border = thin_border
                wb[formatted_name]['G28'].border = thin_border
                wb[formatted_name]['A29'].border = thin_border
                wb[formatted_name]['B29'].border = thin_border
                wb[formatted_name]['C29'].border = thin_border
                wb[formatted_name]['D29'].border = thin_border
                wb[formatted_name]['E29'].border = thin_border
                wb[formatted_name]['F29'].border = thin_border
                wb[formatted_name]['G29'].border = thin_border
                wb[formatted_name]['A30'].border = thin_border
                wb[formatted_name]['B30'].border = thin_border
                wb[formatted_name]['C30'].border = thin_border
                wb[formatted_name]['D30'].border = thin_border
                wb[formatted_name]['E30'].border = thin_border
                wb[formatted_name]['F30'].border = thin_border
                wb[formatted_name]['G30'].border = thin_border
                wb[formatted_name]['A31'].border = thin_border
                wb[formatted_name]['B31'].border = thin_border
                wb[formatted_name]['C31'].border = thin_border
                wb[formatted_name]['D31'].border = thin_border
                wb[formatted_name]['E31'].border = thin_border
                wb[formatted_name]['F31'].border = thin_border
                wb[formatted_name]['G31'].border = thin_border
                wb[formatted_name]['A32'].border = thin_border
                wb[formatted_name]['B32'].border = thin_border
                wb[formatted_name]['C32'].border = thin_border
                wb[formatted_name]['D32'].border = thin_border
                wb[formatted_name]['E32'].border = thin_border
                wb[formatted_name]['F32'].border = thin_border
                wb[formatted_name]['G32'].border = thin_border
                wb[formatted_name]['A33'].border = thin_border
                wb[formatted_name]['B33'].border = thin_border
                wb[formatted_name]['C33'].border = thin_border
                wb[formatted_name]['D33'].border = thin_border
                wb[formatted_name]['E33'].border = thin_border
                wb[formatted_name]['F33'].border = thin_border
                wb[formatted_name]['G33'].border = thin_border
                wb[formatted_name]['A34'].border = thin_border
                wb[formatted_name]['B34'].border = thin_border
                wb[formatted_name]['C34'].border = thin_border
                wb[formatted_name]['D34'].border = thin_border
                wb[formatted_name]['E34'].border = thin_border
                wb[formatted_name]['F34'].border = thin_border
                wb[formatted_name]['G34'].border = thin_border
                wb[formatted_name]['A35'].border = thin_border
                wb[formatted_name]['B35'].border = thin_border
                wb[formatted_name]['C35'].border = thin_border
                wb[formatted_name]['D35'].border = thin_border
                wb[formatted_name]['E35'].border = thin_border
                wb[formatted_name]['F35'].border = thin_border
                wb[formatted_name]['G35'].border = thin_border
                wb[formatted_name]['A36'].border = thin_border
                wb[formatted_name]['B36'].border = thin_border
                wb[formatted_name]['C36'].border = thin_border
                wb[formatted_name]['D36'].border = thin_border
                wb[formatted_name]['E36'].border = thin_border
                wb[formatted_name]['F36'].border = thin_border
                wb[formatted_name]['G36'].border = thin_border
                wb[formatted_name]['A37'].border = thin_border
                wb[formatted_name]['B37'].border = thin_border
                wb[formatted_name]['C37'].border = thin_border
                wb[formatted_name]['D37'].border = thin_border
                wb[formatted_name]['E37'].border = thin_border
                wb[formatted_name]['F37'].border = thin_border
                wb[formatted_name]['G37'].border = thin_border
                wb[formatted_name]['A38'].border = thin_border
                wb[formatted_name]['B38'].border = thin_border
                wb[formatted_name]['C38'].border = thin_border
                wb[formatted_name]['D38'].border = thin_border
                wb[formatted_name]['E38'].border = thin_border
                wb[formatted_name]['F38'].border = thin_border
                wb[formatted_name]['G38'].border = thin_border
                wb[formatted_name].merge_cells('A1:G1')
                a39 = wb[formatted_name]['A39'].alignment = Alignment(horizontal='center')
                a39 = wb[formatted_name]['A39'].alignment = Alignment(wrapText=True)
                wb[formatted_name].page_setup.LeftMargin = 0.75
                wb[formatted_name].page_setup.RightMargin = 0.75
                wb[formatted_name].page_setup.TopMargin = .25
                wb[formatted_name].page_setup.BottomMargin = .25
                wb[formatted_name].page_setup.HeaderMargin = 0.5
                wb[formatted_name].page_setup.FooterMargin = 0.5
                wb[formatted_name].page_setup.fitToPage = True
        wb.save('My Name Application.xlsx')
        messagebox.showinfo("Update Spreadsheet", "The spreadsheet has been updated")



def importpaneltextfile():
        print ("Opening and closing the file.")
        filename = fd.askopenfilename(initialdir="/", title="Select file",
                                      filetypes=(("txt files", "*.txt"), ("all files", "*.*")))
        with open(filename, 'r') as file:
            [panellistbox.insert(1, item) for item in file.readlines()]

def panel1update():
            PanelCombo1['values'] = ['jjjj']


def importapptextfile():
        applicantlistbox.delete(0,END)
        print ("Opening and closing the file.")
        filename = fd.askopenfilename(initialdir="/", title="Select file",
                                      filetypes=(("txt files", "*.txt"), ("all files", "*.*")))
        with open(filename, 'r') as file:
            [applicantlistbox.insert(1, item) for item in file.readlines()]
            Button(page2, text='Import Applicants', activebackground="white", bd=3,bg="white",width=13, command=importapptextfile).grid(row=2, column=2, sticky=W, pady=4)
            

def PDF():
        email_list_box.delete(0,END)
        print ("Opening and closing the pdf file.")
        filename = fd.askopenfilename(initialdir="/", title="Select file",
                                      filetypes=(("pdf files", "*.pdf"), ("all files", "*.*")))
        with open(filename, 'r') as file:
            [email_list_box.insert(1, item) for item in file.readlines()]
            Button(page1, text='Import PDF', activebackground="white", bd=3,bg="white",width=13, command=PDF).grid(row=2, column=3, sticky=W, pady=4)


def printpanel(evt):
        value=str((panellistbox.get(panellistbox.curselection())))
        print (value)         

def printapplicants(evt):
        value=str((applicantlistbox.get(applicantlistbox.curselection())))
        print (value)



def UpdatePage2():
        wb = openpyxl.load_workbook('My Name Application.xlsx')
        wb["Page 2"]['A5'] = "Position: " + str(myjobtitle.get())
        wb["Page 2"]['A6'] = "Cert# " + str(mycert.get())
        wb["Page 2"]['B8'] = "EXPERIENCE-" + str(experience.get())
        wb["Page 2"]['F8'] = "EDUCATION-" + str(education.get())
        wb["Page 2"]['J8'] = "TRAINING-" + str(training.get())
        wb["Page 2"]['N8'] = "AWARDS-" + str(awards.get())
        wb["Page 2"]['R8'] = "INTERVIEW QUESTIONS-" + str(maxinterview.get())
        wb["Page 2"]['A9'] = "Panel Initials"
        row_count = wb["Page 2"].max_row
        print (row_count)
        wb["Page 2"]['A' + row_count] = "SIGNATURE"
        wb.save('My Name Application.xlsx')
        messagebox.showinfo("Update Spreadsheet", "The spreadsheet has been updated")


        

#ws.max_row
#IncludeZeros

##Sheets("Page 2").Range("A" & applicantlastrow + 4).Value = "SIGNATURE"
##Sheets("Page 2").Range("F" & applicantlastrow + 4).Value = "SIGNATURE"
##Sheets("Page 2").Range("P" & applicantlastrow + 4).Value = "SIGNATURE"
##
##Sheets("Page 2").Range("A" & applicantlastrow + 4).Font.Bold = True
##Sheets("Page 2").Range("F" & applicantlastrow + 4).Font.Bold = True
##Sheets("Page 2").Range("P" & applicantlastrow + 4).Font.Bold = True
##Sheets("Page 2").Range("A" & applicantlastrow + 4).Font.Size = 12
##Sheets("Page 2").Range("F" & applicantlastrow + 4).Font.Size = 12
##Sheets("Page 2").Range("P" & applicantlastrow + 4).Font.Size = 12

        wb.save('Applicant Application.xlsx')
        messagebox.showinfo("Update Spreadsheet", "The spreadsheet has been updated")

def UpdateSpreadsheet():
        wb = openpyxl.load_workbook('My Name Application.xlsx')
        wb["Sheet1"]['A1'] = "Your Name: " + str(myjobtitle.get())
        wb.save('My Name Application.xlsx')
        messagebox.showinfo("Update Spreadsheet", "The spreadsheet has been updated")


def chairperson():
    if chairperson_checkbox.get():
        Label(page3, text="Select the Panel Member Titles", bg="lightgray", width=25).grid(row=1, column=5, sticky=W)
        Label(page3, text="ChairPerson", fg="red").grid(row=2, column=5, sticky=W)
    else:
        Label(page3, text="", bg="lightgray", width=10).grid(row=2, column=5, sticky=W)
        
def chairperson_eeo():
    if chairperson_eeo_checkbox.get():
        Label(page3, text="Select the Panel Member Titles", bg="lightgray", width=25).grid(row=1, column=5, sticky=W)
        Label(page3, text="ChairPerson/EEO", fg="red").grid(row=2, column=5, sticky=W)
    else:
        Label(page3, text="", bg="lightgray", width=13).grid(row=2, column=5, sticky=W)

def sme_eeo():
    if sme_eeo_checkbox.get():
        Label(page3, text="Select the Panel Member Titles", bg="lightgray", width=25).grid(row=1, column=5, sticky=W)
        Label(page3, text="SME/EEO", fg="red").grid(row=4, column=5, sticky=W)
    else:
        Label(page3, text="", bg="lightgray", width=7).grid(row=5, column=5, sticky=W)
    


def sme_one():
    if sme_one_checkbox.get():
        Label(page3, text="Select the Panel Member Titles", bg="lightgray", width=25).grid(row=1, column=5, sticky=W)
        Label(page3, text="SME", fg="red").grid(row=4, column=5, sticky=W)
    else:
        Label(page3, text="", bg="lightgray", width=4).grid(row=4, column=5, sticky=W)
    

def sme_two():
    if sme_two_checkbox.get():
        Label(page3, text="Select the Panel Member Titles", bg="lightgray", width=25).grid(row=1, column=5, sticky=W)
        Label(page3, text="SME", fg="red").grid(row=6, column=5, sticky=W)
    else:
        Label(page3, text="", bg="lightgray", width=4).grid(row=6, column=5, sticky=W)
    

def sme_three():
    if sme_three_checkbox.get():
        Label(page3, text="Select the Panel Member Titles", bg="lightgray", width=25).grid(row=1, column=5, sticky=W)
        Label(page3, text="SME", fg="red").grid(row=8, column=5, sticky=W)
    else:
        Label(page3, text="", bg="lightgray", width=4).grid(row=8, column=5, sticky=W)

def fmiddle():
    if fmiddle_checkbox.get():
        messagebox.showinfo("Initials", "Type First and Middle Initial")
        first_first_init_entry_box = Entry(page3, textvariable=ffint, bd=3,bg="red",width=5).grid(row=3, column=8, sticky=W, pady=4)


def smiddle():
    if smiddle_checkbox.get():
        messagebox.showinfo("Initials", "Type First and Middle Initial")
        second_first_init_entry_box = Entry(page3, textvariable=sfint, bd=3,bg="red",width=5).grid(row=5, column=8, sticky=W, pady=4)

def tmiddle():
    if tmiddle_checkbox.get():
        messagebox.showinfo("Initials", "Type First and Middle Initial")
        third_first_init_entry_box = Entry(page3, textvariable=tfint, bd=3,bg="red",width=5).grid(row=7, column=8, sticky=W, pady=4)

def fomiddle():
    if fomiddle_checkbox.get():
        messagebox.showinfo("Initials", "Type First and Middle Initial")
        forth_first_init_entry_box = Entry(page3, textvariable=fofint, bd=3,bg="red",width=5).grid(row=9, column=8, sticky=W, pady=4)


def show():
        numofquest = numofintvquestions.get()

        if numofquest == "6":
            Label(page4, text="Enter the Interview Question Six Points", fg="red").grid(row=14, column=4, sticky=W)
            quest_six=Entry(page4, textvariable=question_six, bd=3,bg="white",width=7).grid(row=15, column=4, sticky=W, pady=4)

        if InterviewYes_checkbox.get():
            Label(page4, text="Enter the Interview Question Two Points", fg="red").grid(row=6, column=4, sticky=W)
            Label(page4, text="Enter the Interview Question Three Points", fg="red").grid(row=8, column=4, sticky=W)
            Label(page4, text="Enter the Interview Question Four Points", fg="red").grid(row=10, column=4, sticky=W)
            Label(page4, text="Enter the Interview Question Five Points", fg="red").grid(row=12, column=4, sticky=W)
            quest_two=Entry(page4, textvariable=question_one, bd=3,bg="white",width=7).grid(row=7, column=4, sticky=W, pady=4)
            quest_three=Entry(page4, textvariable=question_one, bd=3,bg="white",width=7).grid(row=9, column=4, sticky=W, pady=4)
            quest_four=Entry(page4, textvariable=question_one, bd=3,bg="white",width=7).grid(row=11, column=4, sticky=W, pady=4)
            quest_five=Entry(page4, textvariable=question_one, bd=3,bg="white",width=7).grid(row=13, column=4, sticky=W, pady=4)
            

        if InterviewNo_checkbox.get():
            Label(page4, text="Enter the Interview Question Two Points", fg="red").grid(row=6, column=4, sticky=W)
            Label(page4, text="Enter the Interview Question Three Points", fg="red").grid(row=8, column=4, sticky=W)
            Label(page4, text="Enter the Interview Question Four Points", fg="red").grid(row=10, column=4, sticky=W)
            Label(page4, text="Enter the Interview Question Five Points", fg="red").grid(row=12, column=4, sticky=W)
            quest_two=Entry(page4, textvariable=question_two, bd=3,bg="white",width=7).grid(row=7, column=4, sticky=W, pady=4)
            quest_three=Entry(page4, textvariable=question_three, bd=3,bg="white",width=7).grid(row=9, column=4, sticky=W, pady=4)
            quest_four=Entry(page4, textvariable=question_four, bd=3,bg="white",width=7).grid(row=11, column=4, sticky=W, pady=4)
            quest_five=Entry(page4, textvariable=question_five, bd=3,bg="white",width=7).grid(row=13, column=4, sticky=W, pady=4)
            
def printworksheets():
    if survey_checkbox.get():
        messagebox.showinfo("Survey Spreadsheet", "The survey spreadsheets are being printed")
    if score_checkbox.get():
        messagebox.showinfo("Individual Score Spreadsheet", "The individual score spreadsheets are being printed")
    if indnotes_checkbox.get():
        messagebox.showinfo("Individual Notes Spreadsheet", "The individual notes spreadsheet are being printed")
    if intnotes_checkbox.get():
        messagebox.showinfo("Interview Notes Spreadsheet", "The interview notes spreadsheet are being printed")




root = tk.Tk()
root.title("Survey of Interest Application")
root.geometry("800x520+0+0")
root.configure(background='red')
nb = ttk.Notebook(root)

myjobtitle = StringVar()
myemail = StringVar()
myseries = StringVar()
myrpa = StringVar()
mycert = StringVar()
myduedate = StringVar()
mytime = StringVar()
myapplicant = StringVar()
survey_checkbox = IntVar()
score_checkbox = IntVar()
indnotes_checkbox = IntVar()
indnotes_checkbox = IntVar()
intnotes_checkbox = IntVar()
intnotes_checkbox = IntVar()
panel_member1 = StringVar()
panel_member2 = StringVar()
panel_member3 = StringVar()
panel_member4 = StringVar()
ffint = StringVar()
flint = StringVar()
sfint = StringVar()
slint = StringVar()
tfint = StringVar()
tlint = StringVar()
fofint = StringVar()
folint = StringVar()
maxscore = StringVar()
experience = StringVar()
education = StringVar()
training = StringVar()
awards = StringVar()
maxinterview = StringVar()
question_one = StringVar()
question_two = StringVar()
question_three = StringVar()
question_four = StringVar()
question_five = StringVar()
question_six = StringVar()
applicant = StringVar()
numofintvquestions = StringVar()





page1 = ttk.Frame(nb)
Label(page1, text="Enter the Job Title", fg="red").grid(row=1, column=3, sticky=W)
job_title_entry_box = Entry(page1, textvariable=myjobtitle, bd=3,bg="white",width=45).grid(row=2, column=3, sticky=W, pady=4)
Label(page1, text="Enter the Series", fg="red").grid(row=3, column=3, sticky=W)
series_entry_box = Entry(page1, textvariable=myseries, bd=3,bg="white",width=15).grid(row=4, column=3, sticky=W, pady=4)
Label(page1, text="Enter the RPA", fg="red").grid(row=5, column=3, sticky=W)
rpa_entry_box = Entry(page1, textvariable=myrpa, bd=3,bg="white",width=15).grid(row=6, column=3, sticky=W, pady=4)
Label(page1, text="Enter the Cert", fg="red").grid(row=7, column=3, sticky=W)
cert_entry_box = Entry(page1, textvariable=mycert, bd=3,bg="white",width=15).grid(row=8, column=3, sticky=W, pady=4)
Label(page1, text="Enter the Due Date", fg="red").grid(row=9, column=3, sticky=W)
due_date_entry_box = Entry(page1, textvariable=myduedate, bd=3,bg="white",width=15).grid(row=10, column=3, sticky=W, pady=4)
Label(page1, text="Enter the Time", fg="red").grid(row=11, column=3, sticky=W)
time_entry_box = Entry(page1, textvariable=mytime, bd=3,bg="white",width=15).grid(row=12, column=3, sticky=W, pady=4)

Label(page1, text="Email Addresses from PDF", fg="red").grid(row=1, column=1, sticky=W)
email_list_box=Listbox(page1, height=9, width=40)
email_list_box.grid(row=2, column=1, sticky=W, pady=4)
[email_list_box.insert(1, item) for item in (email_finder(pdf_content))]

Button(page1, text='Send Email', activebackground="white", bd=3,bg="white",width=28).grid(row=3, column=4, sticky=W, pady=4)
Button(page1, text='Delete Selected Email Addresses', activebackground="white", bd=3,bg="white",width=30, command=deleteselectedemails).grid(row=5, column=1, sticky=W, pady=4)
email_entry_box = Entry(page1, textvariable=myemail, bd=3,bg="white",width=35).grid(row=6, column=1, sticky=W, pady=4)
Button(page1, text='Add Email Addresses', activebackground="white", bd=3,bg="white",width=30, command=addemails).grid(row=7, column=1, sticky=W, pady=4)


page2 = ttk.Frame(nb)
Label(page2, text="Applicants", fg="black").grid(row=1, column=1, sticky=W)
Label(page2, text="Enter the Job Title", fg="red").grid(row=3, column=1, sticky=W)
job_title_entry_box = Entry(page2, textvariable=myjobtitle, bd=3,bg="white",width=45).grid(row=4, column=1, sticky=W, pady=4)
Label(page2, text="Enter the Series", fg="red").grid(row=5, column=1, sticky=W)
series_entry_box = Entry(page2, textvariable=myseries, bd=3,bg="white",width=15).grid(row=6, column=1, sticky=W, pady=4)
Label(page2, text="Enter the RPA", fg="red").grid(row=7, column=1, sticky=W)
rpa_entry_box = Entry(page2, textvariable=myrpa, bd=3,bg="white",width=15).grid(row=8, column=1, sticky=W, pady=4)
Label(page2, text="Enter the Cert", fg="red").grid(row=9, column=1, sticky=W)
cert_entry_box = Entry(page2, textvariable=mycert, bd=3,bg="white",width=15).grid(row=10, column=1, sticky=W, pady=4)
Checkbutton(page2, text="Further Considering (Declined)").grid(row=3, column=2,sticky=W)
Checkbutton(page2, text="Declined Interview").grid(row=3, column=3,sticky=W)
applicantlistbox=Listbox(page2,width=30,height=8,font=('times',13))
applicantlistbox.bind('<<ListboxSelect>>',printapplicants)
applicantlistbox.place(x=32,y=90)
#Lb1=Listbox(page2, height=8, width=40)
#yscroll = Scrollbar(page2, orient=VERTICAL)
#Lb1['yscrollcommand'] = yscroll.set
#yscroll['command'] = Lb1.yview
applicantlistbox.grid(row=2, column=1, sticky=W, pady=4)
#yscroll.grid(row=2, column=1, rowspan=2, sticky=N+S+E)
Button(page2, text='Import Applicants', activebackground="white", bd=3,fg="red",width=13, command=importapptextfile).grid(row=2, column=2, sticky=W, pady=4)
Button(page2, text='Delete All Applicants', activebackground="white", bd=3,bg="white",width=20, command=deleteallapplicants).grid(row=2, column=3, sticky=W, pady=4)
Button(page2, text='Delete Selected Applicants', activebackground="white", bd=3,bg="white",width=20, command=deleteselectedapplicants).grid(row=2, column=4, sticky=W, pady=4)
applicant_entry_box = Entry(page2, textvariable=myapplicant, bd=3,bg="white",width=35).grid(row=1, column=2, sticky=W, pady=4)
Button(page2, text='Add Applicant', activebackground="white", bd=3,bg="white",width=12, command=addapplicants).grid(row=1, column=3, sticky=W, pady=4)
Button(page2, text='Update', activebackground="white", bd=3,bg="white",width=12).grid(row=4, column=4, sticky=W, pady=4)

page3 = ttk.Frame(nb)

Label(page3, text="**Complete the Items in Red**", fg="red").grid(row=1, column=1, sticky=W)


#Lb13=Listbox(page3,height=6, width=40)
#Lb13.grid(row=3, column=1, sticky=W, pady=4)

Label(page3, text="Enter the # of Panel Members", fg="red").grid(row=9, column=1, sticky=W)
Spinbox1=Spinbox(page3, from_=3,to=4, width=5).grid(row=10, column=1, sticky=W, pady=4)
Label(page3, text="Select the Panel Member Titles", fg="red").grid(row=11, column=1, sticky=W)

panellistbox=Listbox(page3,width=21,height=5, font=('times',13))
panellistbox.bind('<<ListboxSelect>>',printpanel)
panellistbox.place(x=12,y=15)


    
chairperson_checkbox = IntVar()
Checkbutton(page3, text="ChairPerson", variable=chairperson_checkbox, command=chairperson).grid(row=12, column=1,sticky=W)
chairperson_eeo_checkbox = IntVar()
Checkbutton(page3, text="ChairPerson/EEO", variable=chairperson_eeo_checkbox, command=chairperson_eeo).grid(row=13, column=1,sticky=W)

sme_eeo_checkbox = IntVar()
Checkbutton(page3, text="SME/EEO", variable=sme_eeo_checkbox, command=sme_eeo).grid(row=14, column=1,sticky=W)

sme_one_checkbox = IntVar()
Checkbutton(page3, text="SME", variable=sme_one_checkbox, command=sme_one).grid(row=15, column=1,sticky=W)    

sme_two_checkbox = IntVar()
Checkbutton(page3, text="SME", variable=sme_two_checkbox, command=sme_two).grid(row=16, column=1,sticky=W)

sme_three_checkbox = IntVar()
Checkbutton(page3, text="SME", variable=sme_three_checkbox, command=sme_three).grid(row=17, column=1,sticky=W)

a = panel_member1
#fmiddle_checkbox = a[1:]
b = panel_member2
#smiddle_checkbox = b[1:]
c = panel_member3
#tmiddle_checkbox = c[1:]
d = panel_member4
#fomiddle_checkbox = d[1:]

    
fmiddle_checkbox = IntVar()
Checkbutton(page3, text="Include Middle Initial", variable=fmiddle_checkbox, command=fmiddle).grid(row=3, column=6,sticky=W)
smiddle_checkbox = IntVar()
Checkbutton(page3, text="Include Middle Initial", variable=smiddle_checkbox, command=smiddle).grid(row=5, column=6,sticky=W)
tmiddle_checkbox = IntVar()
Checkbutton(page3, text="Include Middle Initial", variable=tmiddle_checkbox, command=tmiddle).grid(row=7, column=6,sticky=W)
fomiddle_checkbox = IntVar()
Checkbutton(page3, text="Include Middle Initial", variable=fomiddle_checkbox, command=fomiddle).grid(row=9, column=6,sticky=W)


PanelCombo1=ttk.Combobox(page3, textvariable=panel_member1, width=25, postcommand=panel1update).grid(row=3, column=5, sticky=W, pady=4)
PanelCombo2=ttk.Combobox(page3, textvariable=panel_member2, width=25).grid(row=5, column=5, sticky=W, pady=4)
PanelCombo3=ttk.Combobox(page3, textvariable=panel_member3, width=25).grid(row=7, column=5, sticky=W, pady=4)
PanelCombo4=ttk.Combobox(page3, textvariable=panel_member4, width=25).grid(row=9, column=5, sticky=W, pady=4)




first_last_init_entry_box = Entry(page3, textvariable=flint, bd=3,bg="white",width=5).grid(row=3, column=11, sticky=W, pady=4)
Label(page3, text="+", bg="white",width=3).grid(row=3, column=10, sticky=W)


second_last_init_entry_box = Entry(page3, textvariable=slint, bd=3,bg="white",width=5).grid(row=5, column=11, sticky=W, pady=4)
Label(page3, text="+", bg="white",width=3).grid(row=5, column=10, sticky=W)

third_last_init_entry_box = Entry(page3, textvariable=tlint, bd=3,bg="white",width=5).grid(row=7, column=11, sticky=W, pady=4)
Label(page3, text="+", bg="white",width=3).grid(row=7, column=10, sticky=W)

forth_first_init_entry_box = Entry(page3, textvariable=fofint, bd=3,bg="white",width=5).grid(row=9, column=11, sticky=W, pady=4)
Label(page3, text="+", bg="white",width=3).grid(row=9, column=10, sticky=W)
forth_last_init_entry_box = Entry(page3, textvariable=folint, bd=3,bg="white",width=5).grid(row=9, column=11, sticky=W, pady=4)


Button(page3, text='Import Panel Members', activebackground="white", bd=3,fg="red",width=20, command=importpaneltextfile).grid(row=8, column=1, sticky=W, pady=4)
Button(page3, text='Delete Panel Members', activebackground="white", bd=3,fg="red",width=20, command=deleteallpanelmembers).grid(row=9, column=2, sticky=W, pady=4)




page4 = ttk.Frame(nb)
Label(page4, text="**Complete the Items in Red**", fg="red").grid(row=1, column=1, sticky=W)
Label(page4, text="Enter the Experience Points", fg="red").grid(row=2, column=1, sticky=W)
Entry(page4, textvariable=experience, bd=3,bg="white",width=7).grid(row=3, column=1, sticky=W, pady=4)
Label(page4, text="Enter the Education Points", fg="red").grid(row=4, column=1, sticky=W)
Entry(page4, textvariable=education, bd=3,bg="white",width=7).grid(row=5, column=1, sticky=W, pady=4)
Label(page4, text="Enter the Training Points", fg="red").grid(row=6, column=1, sticky=W)
Entry(page4, textvariable=training, bd=3,bg="white",width=7).grid(row=7, column=1, sticky=W, pady=4)
Label(page4, text="Enter the Awards Points", fg="red").grid(row=8, column=1, sticky=W)
Entry(page4, textvariable=awards, bd=3,bg="white",width=7).grid(row=9, column=1, sticky=W, pady=4)
Label(page4, text="Individual Max Points", bg="white").grid(row=10, column=1, sticky=W)
Entry(page4, textvariable=maxscore, bd=3,bg="white",width=7).grid(row=11, column=1, sticky=W, pady=4)

Label(page4, text="Enter the # of Interview Questions", fg="red").grid(row=2, column=4, sticky=W)
Spinbox2=Spinbox(page4, from_=5,to=6, textvariable=numofintvquestions, width=8).grid(row=3, column=4, sticky=W, pady=4)

Label(page4, text="Enter the Interview Question One Points", fg="red").grid(row=4, column=4, sticky=W)
quest_one=Entry(page4, textvariable=question_one, bd=3,bg="white",width=7).grid(row=5, column=4, sticky=W, pady=4)



Button(page4, text='Press Update', activebackground="white", bd=3,bg="white",width=10, command=show).grid(row=3, column=4, sticky=E, pady=1)

Label(page4, text="Are the Interview Points all the same?", fg="red").grid(row=2, column=5, sticky=W)   
InterviewYes_checkbox = IntVar()
Checkbutton(page4, text="Yes", variable=InterviewYes_checkbox, command=show).grid(row=3, column=5,sticky=W)
InterviewNo_checkbox = IntVar()
Checkbutton(page4, text="No", variable=InterviewNo_checkbox, command=show).grid(row=4, column=5,sticky=W)

Label(page4, text="Interview Max Points", fg="red").grid(row=18, column=4, sticky=W)
Entry1=Entry(page4, textvariable=maxinterview, bd=3,bg="white",width=7).grid(row=19, column=4, sticky=W, pady=4)


page5 = ttk.Frame(nb)
Button(page5, text='Create Survey Pages', activebackground="white", bd=3,bg="white",width=28).grid(row=3, column=1, sticky=W, pady=4)
Button(page5, text='Create Individual Score Worksheets', activebackground="white", bd=3,bg="white",width=28, command=scoresheets).grid(row=4, column=1, sticky=W, pady=4)
Button(page5, text='Create Individual Worksheet Notes', activebackground="white", bd=3,bg="white",width=28, command=WorksheetNotes).grid(row=5, column=1, sticky=W, pady=4)
Button(page5, text='Create Individual Interview Pages', activebackground="white", bd=3,bg="white",width=28, command=InterviewNotes).grid(row=6, column=1, sticky=W, pady=4)
Button(page5, text='Update Page 2', activebackground="white", bd=3,bg="white",width=28, command=UpdatePage2).grid(row=7, column=1, sticky=W, pady=4)

survey_checkbox = IntVar()
Checkbutton(page5, text="Survey Pages", variable=survey_checkbox).grid(row=9, column=1,sticky=W)
score_checkbox = IntVar()
Checkbutton(page5, text="Individual Score Sheets", variable=score_checkbox).grid(row=10, column=1,sticky=W)
indnotes_checkbox = IntVar()
Checkbutton(page5, text="Individual Notes Sheets", variable=indnotes_checkbox).grid(row=11, column=1,sticky=W)
intnotes_checkbox = IntVar()
Checkbutton(page5, text="Interview Notes Sheets", variable=intnotes_checkbox).grid(row=12, column=1,sticky=W)
Button(page5, text='Page Setup', activebackground="white", bd=3,bg="white",width=28).grid(row=13, column=1, sticky=W, pady=4)
Button(page5, text='Print', activebackground="white", bd=3,bg="white",width=28, command=printworksheets).grid(row=14, column=1, sticky=W, pady=4)




nb.add(page1, text='Email Addresses')
nb.add(page2, text='Job Title')
nb.add(page3, text='Panel Members')
nb.add(page4, text='Applicant/Interview Points')
nb.add(page5, text='Produce/Print Worksheets')

nb.pack(expand=1, fill="both")

root.mainloop()
